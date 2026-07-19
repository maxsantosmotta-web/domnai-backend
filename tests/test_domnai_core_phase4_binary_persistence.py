from contextlib import contextmanager
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.domnai_core import (
    ArtifactGenerationAuthorization,
    ArtifactService,
    BinaryArtifactService,
    InMemoryArtifactStore,
    PostgresArtifactStore,
)


def _sqlite_scope_factory():
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    with engine.begin() as connection:
        connection.execute(text("""
            CREATE TABLE domnai_core_artifacts (
                artifact_id VARCHAR(64) PRIMARY KEY,
                name VARCHAR(500) NOT NULL,
                mime_type VARCHAR(255) NOT NULL,
                content BLOB NOT NULL,
                origin VARCHAR(20) NOT NULL,
                owner_id VARCHAR(255) NOT NULL DEFAULT '',
                metadata_json TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL
            )
        """))

    @contextmanager
    def scope():
        session = session_factory()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    return engine, scope


def test_binary_generation_requires_explicit_or_contextual_authorization():
    service = BinaryArtifactService(ArtifactService(InMemoryArtifactStore()))

    with pytest.raises(PermissionError):
        service.generate_pdf(
            name="relatorio.pdf",
            text="Conteúdo",
            authorization=ArtifactGenerationAuthorization(),
        )


def test_pdf_generation_records_authorization_and_valid_signature():
    artifacts = ArtifactService(InMemoryArtifactStore())
    service = BinaryArtifactService(artifacts)

    artifact = service.generate_pdf(
        name="relatorio.pdf",
        text="Título\n\nConteúdo seguro do relatório.",
        owner_id="user-1",
        authorization=ArtifactGenerationAuthorization(
            explicitly_requested=True,
            source="turn-123",
        ),
    )

    assert artifact.content.startswith(b"%PDF")
    assert artifact.mime_type == "application/pdf"
    assert artifact.metadata["generation_authorized"] is True
    assert artifact.metadata["authorization_mode"] == "explicit_request"
    assert artifact.metadata["authorization_source"] == "turn-123"
    assert artifacts.get(artifact.artifact_id, owner_id="user-1").sha256 == artifact.sha256


def test_xlsx_generation_creates_readable_workbook():
    service = BinaryArtifactService(ArtifactService(InMemoryArtifactStore()))

    artifact = service.generate_xlsx(
        name="dados.xlsx",
        rows=[{"nome": "Max", "valor": 10}, {"nome": "Nathália", "valor": 20}],
        authorization=ArtifactGenerationAuthorization(contextually_accepted=True),
    )

    workbook = load_workbook(BytesIO(artifact.content), read_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    assert values[0] == ("nome", "valor")
    assert values[1] == ("Max", 10)
    assert values[2] == ("Nathália", 20)
    assert artifact.metadata["authorization_mode"] == "contextual_acceptance"


def test_direct_binary_registration_without_authorization_is_blocked():
    artifacts = ArtifactService(InMemoryArtifactStore())

    with pytest.raises(PermissionError):
        artifacts.register_generated_binary(
            name="fraude.pdf",
            mime_type="application/pdf",
            content=b"%PDF-fake",
        )


def test_expired_artifacts_are_hidden_from_get_and_listing():
    store = InMemoryArtifactStore()
    artifacts = ArtifactService(store, now_provider=lambda: 200.0)
    artifact = artifacts.generate_text(
        name="temporario.txt",
        text="conteúdo",
        metadata={"expires_at": 100.0},
    )

    with pytest.raises(KeyError):
        artifacts.get(artifact.artifact_id)
    assert artifacts.list_summaries() == ()


def test_postgres_artifact_store_round_trip_and_owner_filter():
    _, scope = _sqlite_scope_factory()
    store = PostgresArtifactStore(scope)
    artifacts = ArtifactService(store)

    first = artifacts.generate_text(
        name="a.txt",
        text="A",
        owner_id="user-a",
    )
    artifacts.generate_text(
        name="b.txt",
        text="B",
        owner_id="user-b",
    )

    loaded = store.get(first.artifact_id)
    assert loaded is not None
    assert loaded.content == b"A"
    assert loaded.metadata["owner_id"] == "user-a"
    assert [item.name for item in store.list(owner_id="user-a")] == ["a.txt"]


def test_postgres_artifact_store_preserves_binary_content():
    _, scope = _sqlite_scope_factory()
    store = PostgresArtifactStore(scope)
    artifacts = ArtifactService(store)
    binary = BinaryArtifactService(artifacts)

    created = binary.generate_pdf(
        name="persistido.pdf",
        text="PDF persistido",
        owner_id="user-1",
        authorization=ArtifactGenerationAuthorization(explicitly_requested=True),
    )
    loaded = store.get(created.artifact_id)

    assert loaded is not None
    assert loaded.content == created.content
    assert loaded.sha256 == created.sha256
