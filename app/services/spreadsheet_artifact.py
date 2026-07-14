from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class GeneratedSpreadsheet:
    filename: str
    mime_type: str
    content: bytes


def _safe_filename(value: str, extension: str) -> str:
    base = re.sub(r"[^A-Za-z0-9À-ÿ._-]+", "_", str(value or "planilha").strip())
    base = base.strip("._-") or "planilha"
    return f"{base[:120]}.{extension}"


def _normalize_rows(headers: list[str], rows: list[list[Any]]) -> tuple[list[str], list[list[Any]]]:
    clean_headers = [str(item or "").strip()[:180] for item in headers[:50]]
    if not clean_headers:
        raise ValueError("A planilha precisa ter ao menos uma coluna.")

    width = len(clean_headers)
    clean_rows: list[list[Any]] = []
    for row in rows[:5000]:
        values = list(row[:width])
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        clean_rows.append(values)
    return clean_headers, clean_rows


def generate_xlsx(title: str, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> GeneratedSpreadsheet:
    clean_headers, clean_rows = _normalize_rows(headers, rows)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (str(sheet_name or "Dados").strip() or "Dados")[:31]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(clean_headers))}{max(1, len(clean_rows) + 1)}"

    for column_index, header in enumerate(clean_headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(clean_rows, start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    for column_index, header in enumerate(clean_headers, start=1):
        max_length = len(header)
        for row_index in range(2, min(len(clean_rows) + 2, 302)):
            value = worksheet.cell(row=row_index, column=column_index).value
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 42)

    output = io.BytesIO()
    workbook.save(output)
    return GeneratedSpreadsheet(
        filename=_safe_filename(title, "xlsx"),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.getvalue(),
    )


def generate_csv(title: str, headers: list[str], rows: list[list[Any]]) -> GeneratedSpreadsheet:
    clean_headers, clean_rows = _normalize_rows(headers, rows)
    text = io.StringIO(newline="")
    writer = csv.writer(text)
    writer.writerow(clean_headers)
    writer.writerows(clean_rows)
    content = text.getvalue().encode("utf-8-sig")
    return GeneratedSpreadsheet(
        filename=_safe_filename(title, "csv"),
        mime_type="text/csv; charset=utf-8",
        content=content,
    )
